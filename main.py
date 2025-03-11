# FILEPATH: E:/hello world/Excel-Quotation-Program/main.py

import os
import openpyxl
import xlrd
import xlwt


def analyze_excel(file_name):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_dir, file_name)
    _, file_extension = os.path.splitext(file_path)

    if file_extension.lower() == '.xlsx':
        workbook = openpyxl.load_workbook(file_path)
        sheet_names = workbook.sheetnames
    elif file_extension.lower() == '.xls':
        workbook = xlrd.open_workbook(file_path)
        sheet_names = workbook.sheet_names()
    else:
        print(f"错误: 不支持的文件格式 '{file_extension}'")
        return

    print(f"已读出名字为 '{file_name}' 的文件")

    remaining_sheets = [name for name in sheet_names if "已识别" not in name]

    while remaining_sheets:
        print(f"\n剩余 {len(remaining_sheets)} 个工作表待识别")
        for i, sheet_name in enumerate(remaining_sheets, 1):
            print(f"{i}. {sheet_name}")

        sheet_choice = input("请选择要识别的工作表序号（输入'q'退出）: ")
        if sheet_choice.lower() == 'q':
            break
        try:
            sheet_index = int(sheet_choice) - 1
            if 0 <= sheet_index < len(remaining_sheets):
                selected_sheet_name = remaining_sheets[sheet_index]
                analyze_sheet(workbook, selected_sheet_name, file_extension)

                # 标记工作表为已识别
                new_sheet_name = f"{selected_sheet_name}_已识别"
                if file_extension.lower() == '.xlsx':
                    sheet = workbook[selected_sheet_name]
                    sheet.title = new_sheet_name
                    workbook.save(file_path)
                else:
                    # 对于 .xls 文件，创建一个新的工作簿并复制内容
                    new_workbook = xlwt.Workbook()
                    added_sheet_names = []  # 用于跟踪已添加的工作表名称

                    # 复制所有工作表，包括要识别的工作表
                    for sheet_name in sheet_names:
                        old_sheet = workbook.sheet_by_name(sheet_name)
                        if sheet_name == selected_sheet_name:
                            new_name = new_sheet_name
                        else:
                            new_name = sheet_name

                        # 检查是否存在重复的工作表名称
                        suffix = 1
                        while new_name in added_sheet_names:
                            new_name = f"{new_name}_{suffix}"
                            suffix += 1

                        new_sheet = new_workbook.add_sheet(new_name)
                        added_sheet_names.append(new_name)
                        for row in range(old_sheet.nrows):
                            for col in range(old_sheet.ncols):
                                new_sheet.write(row, col, old_sheet.cell_value(row, col))

                    new_workbook.save(file_path)

                    # 重新加载工作簿以更新 sheet_names
                    workbook = xlrd.open_workbook(file_path)
                    sheet_names = workbook.sheet_names()

                remaining_sheets.remove(selected_sheet_name)
                print(f"\n工作表 '{selected_sheet_name}' 已被标记为已识别")
            else:
                print("无效的序号，请重新选择。")
        except ValueError:
            print("请输入有效的数字。")

    # ... 其余代码保持不变 ...

    if not remaining_sheets:
        print("\n所有工作表已识别完成。")
        mark_file = input("是否将整个Excel文件标记为已识别？(是/否): ")
        if mark_file.lower() in ['是', 'y', 'yes']:
            return True

    return False


def analyze_sheet(workbook, sheet_name, file_extension):
    print(f"\n正在分析工作表: '{sheet_name}'")

    if file_extension.lower() == '.xlsx':
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
    else:
        sheet = workbook.sheet_by_name(sheet_name)
        max_row = sheet.nrows

    for row in range(max_row):
        if file_extension.lower() == '.xlsx':
            cell_value = sheet.cell(row=row + 1, column=2).value
        else:
            cell_value = sheet.cell_value(row, 1)  # xlrd 使用 0-based 索引

        if cell_value:
            content = str(cell_value).replace(" ", "")
            char_count = len(content)
            print(f"已识别 B{row + 1} 单元格存在 {char_count} 个字符（已删除空格）: '{content}'")

    print(f"\n工作表 '{sheet_name}' 分析完成")


def get_excel_files():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    excel_files = [f for f in os.listdir(current_dir)
                   if f.endswith(('.xlsx', '.xls')) and "已识别" not in f]
    return excel_files


# 主程序
if __name__ == "__main__":
    while True:
        excel_files = get_excel_files()

        if not excel_files:
            print("当前目录下没有找到未识别的Excel文件，程序将自动退出。")
            break

        print("\n检测到以下未识别的Excel文件：")
        for i, file in enumerate(excel_files, 1):
            print(f"{i}. {file}")

        choice = input("请选择要识别的文件序号（输入'q'退出）: ")
        if choice.lower() == 'q':
            break
        try:
            file_index = int(choice) - 1
            if 0 <= file_index < len(excel_files):
                file_to_analyze = excel_files[file_index]
                should_rename = analyze_excel(file_to_analyze)
                if should_rename:
                    # 将已识别的文件重命名，添加 "_已识别" 后缀
                    new_file_name = os.path.splitext(file_to_analyze)[0] + "_已识别" + \
                                    os.path.splitext(file_to_analyze)[1]
                    os.rename(file_to_analyze, new_file_name)
                    print(f"\n文件 '{file_to_analyze}' 已重命名为 '{new_file_name}'")
            else:
                print("无效的序号，请重新选择。")
        except ValueError:
            print("请输入有效的数字。")

    print("程序已退出。")